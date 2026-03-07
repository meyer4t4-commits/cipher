"""
Data Analysis Agent - Real data processing, analysis, SQL queries, and visualization.
ALL REAL OPERATIONS — actual file reads, real SQL, real pandas. No mock data.
"""

import csv
import json
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from app.agents.base import BaseAgent
from app.agents.models import AgentCapability, AgentResult, AgentTask
from app.core.logging import logger


class DataAgent(BaseAgent):
    """Data analysis, SQL queries, and visualization — all real operations."""

    def __init__(self):
        """Initialize the data agent."""
        super().__init__(
            name="data_agent",
            description="Data analysis, SQL, and visualization",
            version="2.0.0",
            capabilities=[
                AgentCapability(
                    name="sql_query",
                    description="Execute SQL queries against database",
                    category="data",
                    timeout_seconds=30,
                ),
                AgentCapability(
                    name="load_csv",
                    description="Load and analyze CSV files",
                    category="data",
                    timeout_seconds=30,
                ),
                AgentCapability(
                    name="load_json",
                    description="Load and process JSON data",
                    category="data",
                    timeout_seconds=30,
                ),
                AgentCapability(
                    name="load_excel",
                    description="Load and analyze Excel files",
                    category="data",
                    timeout_seconds=30,
                ),
                AgentCapability(
                    name="statistical_analysis",
                    description="Perform statistical analysis on loaded data",
                    category="data",
                    timeout_seconds=30,
                ),
                AgentCapability(
                    name="data_cleaning",
                    description="Clean and transform data",
                    category="data",
                    timeout_seconds=30,
                ),
                AgentCapability(
                    name="generate_chart",
                    description="Generate data visualizations",
                    category="data",
                    timeout_seconds=60,
                ),
                AgentCapability(
                    name="generate_report",
                    description="Generate analysis report",
                    category="data",
                    timeout_seconds=60,
                ),
            ],
        )
        # In-memory data store for loaded datasets (keyed by data_source name)
        self._loaded_data: dict[str, dict] = {}

    async def validate(self, task: AgentTask) -> bool:
        """Validate data task."""
        if not await super().validate(task):
            return False

        operation = task.params.get("operation", "sql_query")

        if operation == "sql_query":
            if "query" not in task.params:
                logger.warning(f"Task {task.task_id}: Missing 'query' parameter")
                return False

        elif operation in ("load_csv", "load_json", "load_excel"):
            if "path" not in task.params:
                logger.warning(f"Task {task.task_id}: Missing 'path' parameter")
                return False

        return True

    async def execute(self, task: AgentTask) -> AgentResult:
        """Execute data operation."""
        operation = task.params.get("operation", "sql_query")

        try:
            if operation == "sql_query":
                return await self._sql_query(task)
            elif operation == "load_csv":
                return await self._load_csv(task)
            elif operation == "load_json":
                return await self._load_json(task)
            elif operation == "load_excel":
                return await self._load_excel(task)
            elif operation == "statistical_analysis":
                return await self._statistical_analysis(task)
            elif operation == "data_cleaning":
                return await self._data_cleaning(task)
            elif operation == "generate_chart":
                return await self._generate_chart(task)
            elif operation == "generate_report":
                return await self._generate_report(task)
            else:
                return AgentResult(
                    task_id=task.task_id,
                    agent_name=self.name,
                    success=False,
                    error=f"Unknown operation: {operation}",
                )
        except Exception as e:
            logger.error(f"Data operation failed: {e}")
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=str(e),
            )

    async def _sql_query(self, task: AgentTask) -> AgentResult:
        """Execute SQL query against Cipher's database — real SQLAlchemy execution."""
        query = task.params.get("query")
        database = task.params.get("database", "default")

        logger.info(f"Executing SQL query on {database}")

        # Safety: block destructive queries unless explicitly allowed
        query_upper = query.strip().upper()
        if any(kw in query_upper for kw in ["DROP", "DELETE", "TRUNCATE", "ALTER"]):
            if not task.params.get("allow_destructive", False):
                return AgentResult(
                    task_id=task.task_id,
                    agent_name=self.name,
                    success=False,
                    error="Destructive SQL queries require explicit allow_destructive=True",
                )

        try:
            from sqlalchemy import text
            from app.db.database import get_engine

            engine = get_engine()
            start = time.time()

            with engine.connect() as conn:
                result = conn.execute(text(query))
                elapsed = (time.time() - start) * 1000

                # For SELECT queries, fetch results
                if query_upper.startswith("SELECT"):
                    columns = list(result.keys()) if result.keys() else []
                    rows = [dict(row._mapping) for row in result.fetchall()]

                    output = {
                        "operation": "sql_query",
                        "database": database,
                        "query": query[:200],
                        "rows_returned": len(rows),
                        "execution_time_ms": round(elapsed, 1),
                        "columns": columns,
                        "results": rows[:100],  # Cap at 100 rows for response
                        "truncated": len(rows) > 100,
                    }
                else:
                    # For INSERT/UPDATE, return rowcount
                    conn.commit()
                    output = {
                        "operation": "sql_query",
                        "database": database,
                        "query": query[:200],
                        "rows_affected": result.rowcount,
                        "execution_time_ms": round(elapsed, 1),
                    }

        except ImportError:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error="Database engine not available",
            )
        except Exception as e:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"SQL execution failed: {str(e)}",
            )

        return AgentResult(
            task_id=task.task_id,
            agent_name=self.name,
            success=True,
            output=output,
        )

    async def _load_csv(self, task: AgentTask) -> AgentResult:
        """Load CSV file — real file read with csv module."""
        path = task.params.get("path")
        limit = task.params.get("limit", 100)

        if not os.path.exists(path):
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"File not found: {path}",
            )

        try:
            file_size = os.path.getsize(path)
            rows = []
            columns = []

            with open(path, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                columns = reader.fieldnames or []

                for i, row in enumerate(reader):
                    rows.append(dict(row))
                    if i >= 10000:  # Safety cap at 10k rows in memory
                        break

            total_rows = len(rows)

            # Store in memory for subsequent analysis
            data_name = Path(path).stem
            self._loaded_data[data_name] = {
                "rows": rows,
                "columns": columns,
                "source": path,
                "type": "csv",
            }

            output = {
                "operation": "load_csv",
                "path": path,
                "rows": total_rows,
                "columns": columns,
                "file_size_bytes": file_size,
                "loaded_at": datetime.utcnow().isoformat(),
                "data_name": data_name,
                "preview_rows": rows[:min(5, limit)],
            }

        except UnicodeDecodeError:
            # Try with latin-1 encoding
            with open(path, "r", newline="", encoding="latin-1") as f:
                reader = csv.DictReader(f)
                columns = reader.fieldnames or []
                rows = [dict(row) for i, row in enumerate(reader) if i < 10000]

            data_name = Path(path).stem
            self._loaded_data[data_name] = {
                "rows": rows, "columns": columns, "source": path, "type": "csv",
            }

            output = {
                "operation": "load_csv",
                "path": path,
                "rows": len(rows),
                "columns": columns,
                "encoding": "latin-1",
                "loaded_at": datetime.utcnow().isoformat(),
                "data_name": data_name,
                "preview_rows": rows[:5],
            }
        except Exception as e:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"Failed to load CSV: {str(e)}",
            )

        return AgentResult(
            task_id=task.task_id,
            agent_name=self.name,
            success=True,
            output=output,
        )

    async def _load_json(self, task: AgentTask) -> AgentResult:
        """Load JSON file — real file read."""
        path = task.params.get("path")

        if not os.path.exists(path):
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"File not found: {path}",
            )

        try:
            file_size = os.path.getsize(path)

            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)

            # Determine structure
            if isinstance(data, list):
                structure = "array"
                record_count = len(data)
                keys = list(data[0].keys()) if data and isinstance(data[0], dict) else []
                preview = data[:5]
            elif isinstance(data, dict):
                structure = "object"
                record_count = len(data)
                keys = list(data.keys())
                preview = {k: data[k] for k in list(data.keys())[:5]}
            else:
                structure = type(data).__name__
                record_count = 1
                keys = []
                preview = str(data)[:500]

            data_name = Path(path).stem
            self._loaded_data[data_name] = {
                "data": data,
                "source": path,
                "type": "json",
                "structure": structure,
            }

            output = {
                "operation": "load_json",
                "path": path,
                "records": record_count,
                "file_size_bytes": file_size,
                "keys": keys[:20],
                "loaded_at": datetime.utcnow().isoformat(),
                "structure": structure,
                "data_name": data_name,
                "preview": preview,
            }

        except json.JSONDecodeError as e:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"Invalid JSON: {str(e)}",
            )
        except Exception as e:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"Failed to load JSON: {str(e)}",
            )

        return AgentResult(
            task_id=task.task_id,
            agent_name=self.name,
            success=True,
            output=output,
        )

    async def _load_excel(self, task: AgentTask) -> AgentResult:
        """Load Excel file — real openpyxl read."""
        path = task.params.get("path")
        sheet = task.params.get("sheet")

        if not os.path.exists(path):
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"File not found: {path}",
            )

        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            ws = wb[sheet] if sheet and sheet in sheet_names else wb.active
            actual_sheet = ws.title

            rows = []
            columns = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    columns = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
                else:
                    row_dict = {}
                    for j, val in enumerate(row):
                        col_name = columns[j] if j < len(columns) else f"col_{j}"
                        row_dict[col_name] = val
                    rows.append(row_dict)
                if i >= 10000:
                    break

            wb.close()

            data_name = Path(path).stem
            self._loaded_data[data_name] = {
                "rows": rows,
                "columns": columns,
                "source": path,
                "type": "excel",
            }

            output = {
                "operation": "load_excel",
                "path": path,
                "sheet": actual_sheet,
                "available_sheets": sheet_names,
                "rows": len(rows),
                "columns": columns,
                "file_size_bytes": os.path.getsize(path),
                "loaded_at": datetime.utcnow().isoformat(),
                "data_name": data_name,
                "preview_rows": rows[:5],
            }

        except ImportError:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error="openpyxl not installed. Install with: pip install openpyxl",
            )
        except Exception as e:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"Failed to load Excel: {str(e)}",
            )

        return AgentResult(
            task_id=task.task_id,
            agent_name=self.name,
            success=True,
            output=output,
        )

    async def _statistical_analysis(self, task: AgentTask) -> AgentResult:
        """Perform statistical analysis on loaded data — real calculations."""
        data_name = task.params.get("data_source") or task.params.get("data_name")
        column = task.params.get("column")
        metrics = task.params.get("metrics", ["mean", "median", "std", "min", "max"])

        if data_name not in self._loaded_data:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"No data loaded with name '{data_name}'. Load data first.",
            )

        dataset = self._loaded_data[data_name]
        rows = dataset.get("rows", dataset.get("data", []))

        if not isinstance(rows, list) or not rows:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error="Loaded data is empty or not in tabular format",
            )

        # Extract numeric values from the specified column (or all numeric columns)
        results = {}
        columns_to_analyze = [column] if column else dataset.get("columns", [])

        for col in columns_to_analyze:
            values = []
            for row in rows:
                val = row.get(col)
                if val is not None:
                    try:
                        values.append(float(val))
                    except (ValueError, TypeError):
                        continue

            if not values:
                continue

            col_stats = {"count": len(values)}

            if "mean" in metrics:
                col_stats["mean"] = round(sum(values) / len(values), 4)
            if "median" in metrics:
                sorted_vals = sorted(values)
                n = len(sorted_vals)
                mid = n // 2
                col_stats["median"] = sorted_vals[mid] if n % 2 else round((sorted_vals[mid - 1] + sorted_vals[mid]) / 2, 4)
            if "std" in metrics and len(values) > 1:
                mean = sum(values) / len(values)
                variance = sum((x - mean) ** 2 for x in values) / (len(values) - 1)
                col_stats["std"] = round(variance ** 0.5, 4)
            if "min" in metrics:
                col_stats["min"] = min(values)
            if "max" in metrics:
                col_stats["max"] = max(values)
            if "sum" in metrics:
                col_stats["sum"] = round(sum(values), 4)

            results[col] = col_stats

        output = {
            "operation": "statistical_analysis",
            "data_source": data_name,
            "columns_analyzed": len(results),
            "results": results,
            "analysis_timestamp": datetime.utcnow().isoformat(),
        }

        return AgentResult(
            task_id=task.task_id,
            agent_name=self.name,
            success=True,
            output=output,
        )

    async def _data_cleaning(self, task: AgentTask) -> AgentResult:
        """Clean and transform data — real operations on loaded data."""
        data_name = task.params.get("data_source") or task.params.get("data_name")
        operations = task.params.get("operations", ["dedup", "drop_nulls"])

        if data_name not in self._loaded_data:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"No data loaded with name '{data_name}'. Load data first.",
            )

        dataset = self._loaded_data[data_name]
        rows = dataset.get("rows", [])
        rows_before = len(rows)
        stats = {"rows_before": rows_before}

        if "dedup" in operations:
            # Remove exact duplicate rows
            seen = set()
            deduped = []
            for row in rows:
                key = tuple(sorted(row.items()))
                if key not in seen:
                    seen.add(key)
                    deduped.append(row)
            stats["rows_deduped"] = rows_before - len(deduped)
            rows = deduped

        if "drop_nulls" in operations:
            # Remove rows where all values are None/empty
            before_null = len(rows)
            rows = [r for r in rows if any(v is not None and str(v).strip() != "" for v in r.values())]
            stats["nulls_dropped"] = before_null - len(rows)

        if "strip_whitespace" in operations:
            for row in rows:
                for key in row:
                    if isinstance(row[key], str):
                        row[key] = row[key].strip()

        if "lowercase" in operations:
            for row in rows:
                for key in row:
                    if isinstance(row[key], str):
                        row[key] = row[key].lower()

        # Update stored data
        dataset["rows"] = rows
        stats["rows_after"] = len(rows)
        stats["rows_removed"] = rows_before - len(rows)

        output = {
            "operation": "data_cleaning",
            "data_source": data_name,
            "operations_applied": operations,
            **stats,
            "cleaning_timestamp": datetime.utcnow().isoformat(),
        }

        return AgentResult(
            task_id=task.task_id,
            agent_name=self.name,
            success=True,
            output=output,
        )

    async def _generate_chart(self, task: AgentTask) -> AgentResult:
        """Generate visualization — real matplotlib chart."""
        chart_type = task.params.get("chart_type", "line")
        data_name = task.params.get("data_source") or task.params.get("data_name")
        x_axis = task.params.get("x_axis")
        y_axis = task.params.get("y_axis")
        title = task.params.get("title", "Chart")

        if data_name and data_name not in self._loaded_data:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"No data loaded with name '{data_name}'. Load data first.",
            )

        try:
            import matplotlib
            matplotlib.use("Agg")  # Non-interactive backend
            import matplotlib.pyplot as plt

            dataset = self._loaded_data.get(data_name, {})
            rows = dataset.get("rows", [])

            if not rows or not x_axis or not y_axis:
                return AgentResult(
                    task_id=task.task_id,
                    agent_name=self.name,
                    success=False,
                    error="Need data_name, x_axis, and y_axis parameters",
                )

            x_vals = []
            y_vals = []
            for row in rows:
                try:
                    x_vals.append(row.get(x_axis, ""))
                    y_vals.append(float(row.get(y_axis, 0)))
                except (ValueError, TypeError):
                    continue

            fig, ax = plt.subplots(figsize=(10, 6))

            if chart_type == "bar":
                ax.bar(range(len(x_vals[:50])), y_vals[:50])
                ax.set_xticks(range(len(x_vals[:50])))
                ax.set_xticklabels(x_vals[:50], rotation=45, ha="right")
            elif chart_type == "scatter":
                ax.scatter(range(len(y_vals[:500])), y_vals[:500], alpha=0.6)
            elif chart_type == "histogram":
                ax.hist(y_vals, bins=30, edgecolor="black")
            else:  # line
                ax.plot(y_vals[:500])

            ax.set_title(title)
            ax.set_xlabel(x_axis)
            ax.set_ylabel(y_axis)
            plt.tight_layout()

            # Save to data directory
            charts_dir = Path("./data/charts")
            charts_dir.mkdir(parents=True, exist_ok=True)
            chart_file = charts_dir / f"chart_{int(datetime.utcnow().timestamp())}.png"
            fig.savefig(chart_file, dpi=150)
            plt.close(fig)

            output = {
                "operation": "generate_chart",
                "chart_type": chart_type,
                "chart_file": str(chart_file),
                "data_points": len(y_vals),
                "generated_at": datetime.utcnow().isoformat(),
            }

        except ImportError:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error="matplotlib not installed. Install with: pip install matplotlib",
            )
        except Exception as e:
            return AgentResult(
                task_id=task.task_id,
                agent_name=self.name,
                success=False,
                error=f"Chart generation failed: {str(e)}",
            )

        return AgentResult(
            task_id=task.task_id,
            agent_name=self.name,
            success=True,
            output=output,
        )

    async def _generate_report(self, task: AgentTask) -> AgentResult:
        """Generate analysis report — real markdown report from loaded data."""
        title = task.params.get("title", "Data Analysis Report")
        data_sources = task.params.get("data_sources", [])

        reports_dir = Path("./data/reports")
        reports_dir.mkdir(parents=True, exist_ok=True)
        report_file = reports_dir / f"report_{int(datetime.utcnow().timestamp())}.md"

        sections = [f"# {title}\n", f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}\n"]

        # Summarize each loaded dataset
        sources = data_sources if data_sources else list(self._loaded_data.keys())

        if not sources:
            sections.append("\n## No Data Loaded\nLoad data before generating a report.\n")
        else:
            for source_name in sources:
                if source_name not in self._loaded_data:
                    sections.append(f"\n## {source_name}\n*Data not loaded*\n")
                    continue

                dataset = self._loaded_data[source_name]
                rows = dataset.get("rows", [])
                columns = dataset.get("columns", [])

                sections.append(f"\n## Dataset: {source_name}\n")
                sections.append(f"- **Source:** {dataset.get('source', 'unknown')}\n")
                sections.append(f"- **Type:** {dataset.get('type', 'unknown')}\n")
                sections.append(f"- **Rows:** {len(rows):,}\n")
                sections.append(f"- **Columns:** {', '.join(columns[:20])}\n")

                # Quick stats for numeric columns
                if rows:
                    sections.append("\n### Summary Statistics\n")
                    for col in columns[:10]:
                        values = []
                        for row in rows:
                            try:
                                values.append(float(row.get(col, "")))
                            except (ValueError, TypeError):
                                continue
                        if values:
                            mean = sum(values) / len(values)
                            sections.append(
                                f"- **{col}:** min={min(values):.2f}, "
                                f"max={max(values):.2f}, mean={mean:.2f}, "
                                f"count={len(values)}\n"
                            )

        report_content = "\n".join(sections)
        report_file.write_text(report_content, encoding="utf-8")

        output = {
            "operation": "generate_report",
            "title": title,
            "report_file": str(report_file),
            "file_size_bytes": os.path.getsize(report_file),
            "data_sources_included": len(sources),
            "generated_at": datetime.utcnow().isoformat(),
        }

        return AgentResult(
            task_id=task.task_id,
            agent_name=self.name,
            success=True,
            output=output,
        )

    async def verify(self, result: AgentResult) -> bool:
        """Verify data operation result."""
        if not isinstance(result.output, dict):
            logger.warning(f"Result {result.task_id}: Output is not a dict")
            return False

        if "operation" not in result.output:
            logger.warning(f"Result {result.task_id}: Missing 'operation'")
            return False

        return True
